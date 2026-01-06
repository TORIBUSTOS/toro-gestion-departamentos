import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sqlite3
import os

# Configuración
OUTPUT_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Calculadora_IPC_OIKOS_V2.xlsx")

def crear_calculadora():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Simulador IPC Semestral"

    # Estilos
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Dark Gray
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Colores por Semestre (Visual)
    sem1_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid") # Blue
    sem2_fill = PatternFill(start_color="FCE7F3", end_color="FCE7F3", fill_type="solid") # Pink
    
    input_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Yellow (Editable)
    result_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # Green (Result)
    result_font = Font(bold=True, color="065F46", size=12)

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Deptos (Solo PB_A por ahora como ejemplo, pero estructura para todos)
    deptos = ["PB A (Hugo)", "PB B", "1 A", "1 B"]

    # --- ENCABEZADOS ---
    ws.cell(row=1, column=1, value="Concepto").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    
    col_idx = 2
    for depto in deptos:
        c = ws.cell(row=1, column=col_idx, value=depto)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        col_idx += 1

    row = 2
    
    # 0. MONTO "BASE" JUNIO 2024
    ws.cell(row=row, column=1, value="MONTO BASE (Junio '24)").font = Font(bold=True)
    c = ws.cell(row=row, column=2, value=234000) # PB A
    c.fill = input_fill
    c.number_format = '$ #,##0'
    
    # Otros en 0
    for i in range(1, len(deptos)):
        ws.cell(row=row, column=2+i, value=0).number_format = '$ #,##0'
    
    monto_previo_cell = "B2" # Referencia
    row += 2

    # --- DATOS REALES INDEC (Inflación Mensual) ---
    # Periodo 1: Dic 2024 -> Ajuste usando Jun'24 a Nov'24
    # Junio: 4.6, Jul: 4.0, Ago: 4.2, Sep: 3.5, Oct: 2.7, Nov: 2.4
    ipc_sem1 = [4.6, 4.0, 4.2, 3.5, 2.7, 2.4] 

    # Periodo 2: Junio 2025 -> Ajuste usando Dic'24 a May'25
    # Proyección ~3.0 mensual
    ipc_sem2 = [3.0, 3.0, 3.0, 3.0, 3.0, 3.0]

    # --- BLOQUE 1: AJUSTE DICIEMBRE 2024 ---
    ws.cell(row=row, column=1, value="AJUSTE DICIEMBRE 2024 (Usa Jun-Nov)").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = sem1_fill
    row += 1

    ipc_cells = []
    meses_sem1 = ["Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre"]
    
    for idx, val in enumerate(ipc_sem1):
        ws.cell(row=row, column=1, value=f"IPC {meses_sem1[idx]} (%)")
        c = ws.cell(row=row, column=2, value=val)
        c.fill = input_fill
        c.number_format = '0.0'
        ipc_cells.append(f"B{row}")
        row += 1
    
    # Cálculo Acumulado
    ws.cell(row=row, column=1, value="Variación Acumulada").font = Font(bold=True)
    formula_acum = f"=((1+{ipc_cells[0]}/100)*(1+{ipc_cells[1]}/100)*(1+{ipc_cells[2]}/100)*(1+{ipc_cells[3]}/100)*(1+{ipc_cells[4]}/100)*(1+{ipc_cells[5]}/100)-1)"
    ws.cell(row=row, column=2, value=formula_acum).number_format = '0.00%'
    var_acum_cell = f"B{row}"
    row += 1

    # Resultado Dic 24
    ws.cell(row=row, column=1, value="MONTO DIC 2024").font = result_font
    ws.cell(row=row, column=1).fill = result_fill
    formula_monto = f"={monto_previo_cell}*(1+{var_acum_cell})"
    c = ws.cell(row=row, column=2, value=formula_monto)
    c.font = result_font
    c.fill = result_fill
    c.number_format = '$ #,##0'
    monto_dic24_cell = f"B{row}"
    
    # Validación Usuario (277k vs 315k real??)
    ws.cell(row=row, column=3, value="<< El sistema calcula $288k pero el real fue $315k?").font = Font(italic=True, color="FF0000")
    
    row += 2

    # --- BLOQUE 2: AJUSTE JUNIO 2025 ---
    ws.cell(row=row, column=1, value="AJUSTE JUNIO 2025 (Usa Dic-May)").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = sem2_fill
    row += 1

    ipc_cells_2 = []
    meses_sem2 = ["Diciembre", "Enero", "Febrero", "Marzo", "Abril", "Mayo"]

    for idx, val in enumerate(ipc_sem2):
        ws.cell(row=row, column=1, value=f"IPC {meses_sem2[idx]} (Est.)")
        c = ws.cell(row=row, column=2, value=val)
        c.fill = input_fill
        c.number_format = '0.0'
        ipc_cells_2.append(f"B{row}")
        row += 1

    ws.cell(row=row, column=1, value="Variación Acumulada").font = Font(bold=True)
    formula_acum_2 = f"=((1+{ipc_cells_2[0]}/100)*(1+{ipc_cells_2[1]}/100)*(1+{ipc_cells_2[2]}/100)*(1+{ipc_cells_2[3]}/100)*(1+{ipc_cells_2[4]}/100)*(1+{ipc_cells_2[5]}/100)-1)"
    ws.cell(row=row, column=2, value=formula_acum_2).number_format = '0.00%'
    var_acum_cell_2 = f"B{row}"
    row += 1

    ws.cell(row=row, column=1, value="MONTO JUNIO 2025").font = result_font
    ws.cell(row=row, column=1).fill = result_fill
    formula_monto_2 = f"={monto_dic24_cell}*(1+{var_acum_cell_2})"
    c = ws.cell(row=row, column=2, value=formula_monto_2)
    c.font = result_font
    c.fill = result_fill
    c.number_format = '$ #,##0'

    # Ajustar ancho
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20

    wb.save(OUTPUT_FILE)
    print(f"Calculadora V2 generada: {OUTPUT_FILE}")

if __name__ == "__main__":
    crear_calculadora()
