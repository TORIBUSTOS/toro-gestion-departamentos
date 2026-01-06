import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sqlite3
import os

# Configuración
OUTPUT_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Calculadora_IPC_OIKOS.xlsx")

def crear_calculadora():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calculadora IPC"

    # Estilos
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Dark Gray
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    sub_header_fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid") # Gray
    sub_header_font = Font(color="FFFFFF", bold=True)
    
    input_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Light Yellow (Editable)
    result_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # Light Green (Result)
    result_font = Font(bold=True, color="065F46")

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Obtener Deptos de DB
    conn = sqlite3.connect('toro_gestion.db')
    cursor = conn.cursor()
    cursor.execute("SELECT alias FROM departamentos")
    deptos = [row[0] for row in cursor.fetchall()]
    conn.close()

    # Si no hay deptos, usar genéricos
    if not deptos:
        deptos = ["PB A", "PB B", "1 A", "1 B"]

    # --- ENCABEZADOS ---
    ws.cell(row=1, column=1, value="Concepto").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    
    col_idx = 2
    for depto in deptos:
        c = ws.cell(row=1, column=col_idx, value=depto)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        col_idx += 1

    # --- ESTRUCTURA FILAS ---
    row = 2
    
    # 1. MONTO INICIAL
    ws.cell(row=row, column=1, value="MONTO INICIAL CONTRATO").font = Font(bold=True)
    for i in range(len(deptos)):
        c = ws.cell(row=row, column=2+i)
        c.fill = input_fill
        c.number_format = '$ #,##0'
        # Ejemplo para PB A (Hugo)
        if deptos[i] == "PB_A" or deptos[i] == "PB A":
            c.value = 130000
    monto_anterior_rows = [row] # Guardamos referencia fila monto
    row += 2

    # --- BLOQUES DE AJUSTE (Generamos 4 semestres proyectados) ---
    ajustes = ["1º AJUSTE (Semestre 1)", "2º AJUSTE (Semestre 2)", "3º AJUSTE (Semestre 3)", "4º AJUSTE (Semestre 4)"]
    
    # Datos IPC reales para pre-llenar (Dic23-May24 y Jun24-Nov24)
    ipc_data = [
        [25.5, 20.6, 13.2, 11.0, 8.8, 4.2], # Semestre 1 (Brutal)
        [4.6, 4.0, 4.2, 3.5, 2.7, 2.4],     # Semestre 2 (Bajando)
        [3.0, 3.0, 3.0, 3.0, 3.0, 3.0],     # Proyección Semestre 3
        [2.5, 2.5, 2.5, 2.5, 2.5, 2.5]      # Proyección Semestre 4
    ]

    for idx_ajuste, titulo in enumerate(ajustes):
        # Título Bloque
        ws.cell(row=row, column=1, value=titulo).font = sub_header_font
        ws.cell(row=row, column=1).fill = sub_header_fill
        row += 1

        # Filas de IPC Mensual
        ipc_rows = []
        data_semestre = ipc_data[idx_ajuste] if idx_ajuste < len(ipc_data) else [0]*6

        for mes in range(1, 7):
            label = f"IPC Mes {mes} (%)"
            ws.cell(row=row, column=1, value=label)
            
            for i in range(len(deptos)):
                c = ws.cell(row=row, column=2+i)
                c.fill = input_fill
                c.value = data_semestre[mes-1]
                c.number_format = '0.0'
            
            ipc_rows.append(row)
            row += 1

        # Cálculo Acumulado
        ws.cell(row=row, column=1, value="IPC Acumulado (%)").font = Font(bold=True)
        acum_row = row
        for i in range(len(deptos)):
            col_letter = get_column_letter(2+i)
            # Formula excel: =((1+B5/100)*(1+B6/100)... - 1)
            factors = [f"(1+{col_letter}{r}/100)" for r in ipc_rows]
            formula = f"=({'*'.join(factors)} - 1)"
            
            c = ws.cell(row=row, column=2+i, value=formula)
            c.number_format = '0.00%'
            c.font = Font(bold=True)
        row += 1

        # Nuevo Monto
        ws.cell(row=row, column=1, value="NUEVO MONTO ALQUILER").font = result_font
        ws.cell(row=row, column=1).fill = result_fill
        
        monto_previo_row = monto_anterior_rows[-1]
        
        for i in range(len(deptos)):
            col_letter = get_column_letter(2+i)
            # Formula: = Monto_Anterior * (1 + Acumulado)
            formula = f"={col_letter}{monto_previo_row} * (1 + {col_letter}{acum_row})"
            
            c = ws.cell(row=row, column=2+i, value=formula)
            c.font = result_font
            c.fill = result_fill
            c.number_format = '$ #,##0'
            c.border = border

        monto_anterior_rows.append(row) # Este monto es base para el siguiente
        row += 2 # Espacio

    # Ajustar anchos
    ws.column_dimensions['A'].width = 30
    for i in range(len(deptos)):
        ws.column_dimensions[get_column_letter(2+i)].width = 18

    # Guardar
    wb.save(OUTPUT_FILE)
    print(f"Calculadora creada en: {OUTPUT_FILE}")

if __name__ == "__main__":
    crear_calculadora()
